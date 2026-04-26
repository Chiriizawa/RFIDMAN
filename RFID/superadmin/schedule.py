from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from functools import wraps
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
import os
import io
import openpyxl

load_dotenv()

schedule_bp = Blueprint("schedule", __name__, template_folder="templates")


# ─────────────────────────────────────────────
# AUTH DECORATOR
# ─────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("sadmin_logged_in"):
            flash("Please log in to access the admin portal.", "warning")
            return redirect(url_for("sadmin.login"))
        return f(*args, **kwargs)
    return decorated_function


# ─────────────────────────────────────────────
# DB CONNECTION
# ─────────────────────────────────────────────

def get_db_connection():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise Exception("❌ DATABASE_URL not found in .env")
    return psycopg2.connect(database_url.strip(), sslmode="require")


# ─────────────────────────────────────────────
# SCHEDULE PAGE (MAIN)
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule", methods=["GET"])
@login_required
def schedule():
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        search = request.args.get("search", "").strip()

        query = """
            SELECT
                sc.id,
                sc.day,
                sc.time,
                sc.subject,
                sc.room_id,
                sc.section_id,
                sc.teacher_id,
                r.room_name,
                s.section_name,
                s.year_level,
                CONCAT_WS(' ', t.first_name, t.middle_name, t.last_name) AS teacher
            FROM schedules sc
            LEFT JOIN rooms    r ON sc.room_id    = r.id
            LEFT JOIN sections s ON sc.section_id = s.id
            LEFT JOIN teachers t ON sc.teacher_id = t.id
        """
        params = []

        if search:
            query += """
                WHERE
                    LOWER(COALESCE(sc.subject,      '')) LIKE LOWER(%s)
                    OR LOWER(COALESCE(sc.day,       '')) LIKE LOWER(%s)
                    OR LOWER(COALESCE(r.room_name,  '')) LIKE LOWER(%s)
                    OR LOWER(COALESCE(s.section_name,'')) LIKE LOWER(%s)
            """
            like = f"%{search}%"
            params.extend([like, like, like, like])

        query += " ORDER BY sc.id ASC"
        cur.execute(query, params)
        schedules = cur.fetchall()

        cur.execute("SELECT id, room_name FROM rooms ORDER BY room_name ASC")
        rooms = cur.fetchall()

        cur.execute("SELECT id, section_name, year_level FROM sections ORDER BY section_name ASC")
        sections = cur.fetchall()

        cur.execute("""
            SELECT id, CONCAT_WS(' ', first_name, middle_name, last_name) AS name 
            FROM teachers 
            ORDER BY first_name, last_name ASC
        """)
        teachers = cur.fetchall()

        return render_template(
            "superadmin/schedule.html",
            schedules=schedules,
            rooms=rooms,
            sections=sections,
            teachers=teachers,
            total_schedules=len(schedules),
            search=search,
        )

    except Exception as e:
        flash(f"Database error: {str(e)}", "error")
        return render_template(
            "superadmin/schedule.html",
            schedules=[], rooms=[], sections=[], teachers=[],
            total_schedules=0, search="",
        )
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ─────────────────────────────────────────────
# ADD SCHEDULE
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule/add", methods=["POST"])
@login_required
def add_schedule():
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        section_id = request.form.get("section_id", "").strip()
        day        = request.form.get("day",        "").strip()
        time       = request.form.get("time",       "").strip()   # single combined field
        subject    = request.form.get("subject",    "").strip()
        room_id    = request.form.get("room_id",    "").strip()
        teacher_id = request.form.get("teacher_id", "").strip()

        if not all([section_id, day, time, subject, room_id]):
            flash("All fields are required.", "error")
            return redirect(url_for("schedule.schedule"))

        # Convert teacher_id to int or None
        if teacher_id == "":
            teacher_id = None
        else:
            try:
                teacher_id = int(teacher_id)
            except ValueError:
                teacher_id = None

        cur.execute("""
            INSERT INTO schedules (section_id, day, time, subject, room_id, teacher_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (section_id, day, time, subject, room_id, teacher_id))
        conn.commit()

        flash("Schedule added successfully.", "success")

    except Exception as e:
        if conn: conn.rollback()
        flash(f"Database error: {str(e)}", "error")
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    return redirect(url_for("schedule.schedule"))


# ─────────────────────────────────────────────
# UPDATE SCHEDULE
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule/update/<int:schedule_id>", methods=["POST"])
@login_required
def update_schedule(schedule_id):
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        section_id = request.form.get("section_id", "").strip()
        day        = request.form.get("day",        "").strip()
        time       = request.form.get("time",       "").strip()   # single combined field
        subject    = request.form.get("subject",    "").strip()
        room_id    = request.form.get("room_id",    "").strip()
        teacher_id = request.form.get("teacher_id", "").strip()

        if not all([section_id, day, time, subject, room_id]):
            flash("All fields are required.", "error")
            return redirect(url_for("schedule.schedule"))

        # Convert teacher_id to int or None
        if teacher_id == "":
            teacher_id = None
        else:
            try:
                teacher_id = int(teacher_id)
            except ValueError:
                teacher_id = None

        cur.execute("""
            UPDATE schedules
            SET section_id = %s,
                day        = %s,
                time       = %s,
                subject    = %s,
                room_id    = %s,
                teacher_id = %s
            WHERE id = %s
        """, (section_id, day, time, subject, room_id, teacher_id, schedule_id))
        conn.commit()

        flash("Schedule updated successfully.", "success")

    except Exception as e:
        if conn: conn.rollback()
        flash(f"Database error: {str(e)}", "error")
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    return redirect(url_for("schedule.schedule"))


# ─────────────────────────────────────────────
# DELETE SCHEDULE
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule/delete/<int:schedule_id>", methods=["POST"])
@login_required
def delete_schedule(schedule_id):
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("DELETE FROM schedules WHERE id = %s", (schedule_id,))
        conn.commit()

        flash("Schedule deleted successfully.", "success")

    except Exception as e:
        if conn: conn.rollback()
        flash(f"Database error: {str(e)}", "error")
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    return redirect(url_for("schedule.schedule"))


# ─────────────────────────────────────────────
# ADD ROOM
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule/room/add", methods=["POST"])
@login_required
def add_room():
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        room_name = request.form.get("room_name", "").strip()

        if not room_name:
            flash("Room name is required.", "error")
            return redirect(url_for("schedule.schedule"))

        cur.execute(
            "SELECT id FROM rooms WHERE LOWER(room_name) = LOWER(%s)", (room_name,)
        )
        if cur.fetchone():
            flash(f'Room "{room_name}" already exists.', "error")
            return redirect(url_for("schedule.schedule"))

        cur.execute("INSERT INTO rooms (room_name) VALUES (%s)", (room_name,))
        conn.commit()

        flash(f'Room "{room_name}" added successfully.', "success")

    except Exception as e:
        if conn: conn.rollback()
        flash(f"Database error: {str(e)}", "error")
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    return redirect(url_for("schedule.schedule"))


# ─────────────────────────────────────────────
# DELETE ROOM
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule/room/delete/<int:room_id>", methods=["POST"])
@login_required
def delete_room(room_id):
    conn = None
    cur  = None
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute(
            "SELECT COUNT(*) AS cnt FROM schedules WHERE room_id = %s", (room_id,)
        )
        if cur.fetchone()["cnt"] > 0:
            flash("Cannot delete room — it is used in existing schedules.", "error")
            return redirect(url_for("schedule.schedule"))

        cur.execute("DELETE FROM rooms WHERE id = %s", (room_id,))
        conn.commit()

        flash("Room deleted successfully.", "success")

    except Exception as e:
        if conn: conn.rollback()
        flash(f"Database error: {str(e)}", "error")
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    return redirect(url_for("schedule.schedule"))


# ─────────────────────────────────────────────
# IMPORT SCHEDULE (Excel)
# ─────────────────────────────────────────────

@schedule_bp.route("/schedule/import", methods=["POST"])
@login_required
def import_schedule():
    file = request.files.get("file")

    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "error")
        return redirect(url_for("schedule.schedule"))

    conn     = None
    cur      = None
    imported = 0
    skipped  = 0
    warnings = []

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file.read()), read_only=True, data_only=True
        )
        ws = wb.active

        raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in raw_headers
        ]

        required_cols = {"day", "time", "subject", "section_name", "room_name"}
        missing_cols  = required_cols - set(headers)
        if missing_cols:
            flash(f"Missing required columns: {', '.join(sorted(missing_cols))}", "error")
            return redirect(url_for("schedule.schedule"))

        idx = {h: i for i, h in enumerate(headers)}

        valid_rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            day          = str(row[idx["day"]]          or "").strip()
            time         = str(row[idx["time"]]         or "").strip()
            subject      = str(row[idx["subject"]]      or "").strip()
            section_name = str(row[idx["section_name"]] or "").strip()
            room_name    = str(row[idx["room_name"]]    or "").strip()

            if not all([day, time, subject, section_name, room_name]):
                warnings.append(f"Row {row_num}: missing value(s) — skipped.")
                skipped += 1
                continue

            valid_rows.append({
                "row_num":      row_num,
                "day":          day,
                "time":         time,
                "subject":      subject,
                "section_name": section_name,
                "room_name":    room_name,
            })

        if not valid_rows:
            flash("No valid rows found in the file.", "error")
            return redirect(url_for("schedule.schedule"))

        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=RealDictCursor)

        for r in valid_rows:
            try:
                cur.execute(
                    "SELECT id FROM sections WHERE LOWER(section_name) = LOWER(%s)",
                    (r["section_name"],)
                )
                section_row = cur.fetchone()
                if not section_row:
                    warnings.append(
                        f"Row {r['row_num']}: section '{r['section_name']}' not found — skipped."
                    )
                    skipped += 1
                    continue
                section_id = section_row["id"]

                cur.execute(
                    "SELECT id FROM rooms WHERE LOWER(room_name) = LOWER(%s)",
                    (r["room_name"],)
                )
                room_row = cur.fetchone()
                if room_row:
                    room_id = room_row["id"]
                else:
                    cur.execute(
                        "INSERT INTO rooms (room_name) VALUES (%s) RETURNING id",
                        (r["room_name"],)
                    )
                    room_id = cur.fetchone()["id"]

                # Skip exact duplicates
                cur.execute("""
                    SELECT id FROM schedules
                    WHERE day        = %s
                      AND time       = %s
                      AND section_id = %s
                      AND room_id    = %s
                """, (r["day"], r["time"], section_id, room_id))
                if cur.fetchone():
                    skipped += 1
                    continue

                cur.execute("""
                    INSERT INTO schedules (section_id, day, time, subject, room_id)
                    VALUES (%s, %s, %s, %s, %s)
                """, (section_id, r["day"], r["time"], r["subject"], room_id))

                imported += 1

            except Exception as row_err:
                print(f"Error processing row {r['row_num']}: {str(row_err)}")
                skipped += 1
                continue

        conn.commit()

        msg = f"✅ Successfully imported {imported} schedules. Skipped {skipped} duplicate/invalid rows."
        if warnings:
            msg += " Issues: " + " | ".join(warnings[:5])
            if len(warnings) > 5:
                msg += f" … and {len(warnings) - 5} more."

        flash(msg, "success" if imported > 0 else "error")

    except Exception as e:
        if conn: conn.rollback()
        flash(f"❌ Import failed: {str(e)}", "error")
    finally:
        if cur:  cur.close()
        if conn: conn.close()

    return redirect(url_for("schedule.schedule"))